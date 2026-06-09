import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from final_decision_system import FallRiskDecisionSystem

system = FallRiskDecisionSystem()

input_path = "/Users/szepan/Documents/Master's Study/BIDH5001 & BIDH5002/Code and results/mimic_IV_fall_cases_24.csv"
#input_path = "/Users/szepan/Documents/Master's Study/BIDH5001 & BIDH5002/MIMIC-IV clinical notes cases/mimic_IV_fall_cases_24.csv"

input_file = Path(input_path)
output_dir = input_file.parent

if "random cases" in input_file.stem.lower():
    output_prefix = "random_cases"
else:
    output_prefix = input_file.stem.replace(" ", "_")

output_path = output_dir / f"{output_prefix}_output_with_results.csv"
summary_path = output_dir / f"{output_prefix}_output_risk_level_counts.csv"
recommendations_path = output_dir / f"{output_prefix}_output_recommendations_flat.csv"
manual_template_path = output_dir / "Manual Evaluation Results template.xlsx"
manual_output_path = output_dir / f"{output_prefix}_Manual Evaluation Results_filled.xlsx"

df = pd.read_csv(input_path, encoding="latin-1")
discharge_col = "text"
nonblank_mask = df[discharge_col].fillna("").str.strip().ne("")


def analyse_case(text: str):
    if pd.isna(text):
        return None
    text = str(text)
    if not text.strip():
        return None
    return json.loads(system.analyse_json(text))


def get_recommendations(result):
    if result is None:
        return []
    return result.get("recommendations", [])


def get_extracted_variables(result):
    if result is None:
        return {}
    return result.get("extracted_variables", {})


def get_extraction_status(result):
    if result is None:
        return {}
    return result.get("extraction_status", {})


def get_extraction_evidence(result):
    if result is None:
        return {}
    return result.get("extraction_evidence", {})


def normalise_excel_value(value):
    if value is None:
        return "null"
    if isinstance(value, float) and pd.isna(value):
        return "null"
    return value


def safe_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    for i in range(1, 100):
        candidate = path.with_name(f"{path.stem}_{i}{path.suffix}")
        if not candidate.exists():
            return candidate
    return path


output_path = safe_output_path(output_path)
summary_path = safe_output_path(summary_path)
recommendations_path = safe_output_path(recommendations_path)
manual_output_path = safe_output_path(manual_output_path)


def build_manual_evaluation_sheet(source_df: pd.DataFrame, results_series: pd.Series):
    wb = load_workbook(manual_template_path)
    ws = wb.active

    header_rows = [1, 2]
    data_start_row = 3

    col_map = {}
    for col in range(1, ws.max_column + 1):
        for hr in header_rows:
            value = ws.cell(hr, col).value
            if value:
                col_map.setdefault(str(value).strip(), col)

    alias_map = {
        "functional_disabled": "functional_diabled",
        "poor_gait_balance_muscle_strength": "poor_gait_balance_muslce_strength",
        "is_cognitive_impaired": "is_cognitive_impaired ",
    }

    def write_if_exists(excel_row, col_name, value):
        lookup = alias_map.get(col_name, col_name)
        if lookup in col_map:
            ws.cell(excel_row, col_map[lookup]).value = normalise_excel_value(value)

    for i, (_, row) in enumerate(source_df.iterrows()):
        excel_row = data_start_row + i
        result = results_series.iloc[i]
        extracted_variables = get_extracted_variables(result)

        write_if_exists(excel_row, "note_id", row.get("note_id"))
        write_if_exists(excel_row, "hadm_id", row.get("hadm_id"))

        for var_name, value in extracted_variables.items():
            if var_name.startswith("_"):
                continue
            write_if_exists(excel_row, var_name, value)

    wb.save(manual_output_path)


# discharge summary
results = df[discharge_col].apply(analyse_case)

df["risk_json"] = results.apply(
    lambda result: json.dumps(result, ensure_ascii=False) if result is not None else None
)
df["risk_level"] = results.apply(
    lambda result: result.get("risk_level") if result is not None else None
)
df["condition"] = results.apply(
    lambda result: result.get("condition") if result is not None else None
)
df["summary"] = results.apply(
    lambda result: result.get("summary") if result is not None else None
)
df["extraction_status_json"] = results.apply(
    lambda result: json.dumps(get_extraction_status(result), ensure_ascii=False)
    if result is not None else None
)
df["extraction_evidence_json"] = results.apply(
    lambda result: json.dumps(get_extraction_evidence(result), ensure_ascii=False)
    if result is not None else None
)
df["recommendation_count"] = results.apply(
    lambda result: len(get_recommendations(result))
)
df["recommendation_categories"] = results.apply(
    lambda result: " | ".join(
        rec.get("category", "") for rec in get_recommendations(result)
    )
)
df["recommendation_texts"] = results.apply(
    lambda result: "\n\n---\n\n".join(
        f"[{rec.get('category', '')}] {rec.get('text', '')}"
        for rec in get_recommendations(result)
    )
)
df["recommendation_sources"] = results.apply(
    lambda result: " | ".join(
        rec.get("source", "") for rec in get_recommendations(result)
    )
)

df.to_csv(output_path, index=False)

risk_counts = (
    df.loc[nonblank_mask, "risk_level"]
    .fillna("Analysis failed")
    .value_counts(dropna=False)
    .rename_axis("risk_level")
    .reset_index(name="count")
)

risk_counts.to_csv(summary_path, index=False)

recommendation_rows = []
for idx, result in results.items():
    if result is None:
        continue
    for rec in result.get("recommendations", []):
        recommendation_rows.append({
            "case_index": idx,
            "note_id": df.at[idx, "note_id"] if "note_id" in df.columns else None,
            "hadm_id": df.at[idx, "hadm_id"] if "hadm_id" in df.columns else None,
            "risk_level": result.get("risk_level"),
            "condition": result.get("condition"),
            "summary": result.get("summary"),
            "recommendation_index": rec.get("index"),
            "recommendation_category": rec.get("category"),
            "recommendation_text": rec.get("text"),
            "recommendation_source": rec.get("source"),
            "recommendation_grade": rec.get("grade"),
            "recommendation_is_conditional": rec.get("is_conditional"),
            "recommendation_condition_note": rec.get("condition_note"),
        })

recommendations_df = pd.DataFrame(recommendation_rows)
recommendations_df.to_csv(recommendations_path, index=False)

build_manual_evaluation_sheet(df, results)

print("Risk level counts:")
print(risk_counts.to_string(index=False))
print(f"\nProcessed discharge summaries: {int(nonblank_mask.sum())}")
print(f"Blank discharge summaries skipped: {int((~nonblank_mask).sum())}")
print(f"\nDetailed output saved to: {output_path}")
print(f"Risk count summary saved to: {summary_path}")
print(f"Flattened recommendations saved to: {recommendations_path}")
print(f"Manual Evaluation Results-format workbook saved to: {manual_output_path}")
